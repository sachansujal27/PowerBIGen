from rest_framework.decorators import api_view
from rest_framework.response import Response
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook

from .models import Record
from .serializers import RecordSerializer, BulkRecordCreateSerializer

from django.db import connection


# ----------------------------
# LIST RECORDS
# ----------------------------
@api_view(["GET"])
def list_records(request):
    records = Record.objects.all().order_by("id")
    serializer = RecordSerializer(records, many=True)

    return Response({
        "count": records.count(),
        "results": serializer.data
    })


# ----------------------------
# BULK CREATE
# ----------------------------
@api_view(["POST"])
def bulk_create_records(request):

    serializer = BulkRecordCreateSerializer(data=request.data)

    if serializer.is_valid():

        # Remove previous dataset
        Record.objects.all().delete()

        serializer.save()

        return Response({
            "message": "Dataset saved successfully."
        })

    return Response(serializer.errors, status=400)

# ----------------------------
# EXPORT EXCEL
# ----------------------------
@api_view(["GET"])
def export_excel(request):

    records = Record.objects.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Records"

    if records.exists():

        headers = []

        for record in records:
            for key in record.data.keys():
                if key not in headers:
                    headers.append(key)

        ws.append(headers)

        for record in records:
            ws.append([
                record.data.get(h, "")
                for h in headers
            ])

    from datetime import datetime

    filename = f"records_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response[
        "Content-Disposition"
    ] = f'attachment; filename="{filename}"'

    wb.save(response)

    return response

    # ----------------------------
# IMPORT EXCEL
# ----------------------------
@api_view(["POST"])
def import_excel(request):

    file = request.FILES.get("file")

    if not file:
        return Response({"error": "No file uploaded"}, status=400)

    wb = load_workbook(file, data_only=True)
    ws = wb.active

    # Read headers
    headers = [ str(cell.value).strip()
    for cell in ws[1]
    if cell.value is not None]

    for cell in ws[1]:
        if cell.value is not None:
            headers.append(str(cell.value).strip())

    Record.objects.all().delete()

    objects = []

    for row in ws.iter_rows(min_row=2, values_only=True):

        # Skip empty rows
        if all(value is None for value in row):
            continue

        row_data = {}

        for header, value in zip(headers, row):
            row_data[header] = value

        objects.append(
            Record(data=row_data)
        )

    Record.objects.bulk_create(objects)

    return Response({
        "message": f"{len(objects)} records imported successfully."
    })

# =-------------------
# EDIT DATA
# --------------
@api_view(["PUT"])
def update_record(request, pk):

    try:
        record = Record.objects.get(pk=pk)
    except Record.DoesNotExist:
        return Response({"error": "Record not found"}, status=404)

    record.data = request.data["data"]
    record.save()

    return Response({
        "message": "Record updated successfully."
    })